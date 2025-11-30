import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
# import pythoncom # Pre copy_to_clipboard_excel, ponechávam zakomentované, kým to nebudeš potrebovať
# import win32com.client as win32 # Pre copy_to_clipboard_excel, ponechávam zakomentované
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os

class ExcelFilterApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Excel Filter + Správna hmotnosť")
        self.df = None
        self.filtered_df = None
        self.filter_col = None
        self.current_edit_id = None
        self.filter_value_num = 50
        self.row_colors = {}

        # GUI
        tk.Button(master, text="Vybrať Excel súbor", command=self.open_file).pack(pady=5)
        tk.Label(master, text="Vyber stĺpec pre filter (>= hodnota):").pack()
        self.filter_col_var = tk.StringVar()
        self.filter_col_dropdown = ttk.Combobox(master, textvariable=self.filter_col_var, state="readonly")
        self.filter_col_dropdown.pack(pady=2)

        tk.Label(master, text="Zadaj hodnotu:").pack()
        self.filter_val_entry = tk.Entry(master)
        self.filter_val_entry.insert(0, str(self.filter_value_num))
        self.filter_val_entry.pack(pady=2)

        tk.Button(master, text="Spracovať", command=self.preview_results).pack(pady=5)
        #tk.Button(master, text="Skopírovať do schránky ako Excel", command=self.copy_to_clipboard_excel).pack(pady=5)
        tk.Button(master, text="Uložiť do Excelu", command=self.save_results).pack(pady=5)

        self.tree_frame = tk.Frame(master)
        self.tree_frame.pack(fill=tk.BOTH, expand=True)
        self.tree = None

        # Style
        self.style = ttk.Style()
        self.style.theme_use("default")
        self.style.configure("Treeview", rowheight=22, fieldbackground='lightyellow', foreground='black')
        self.style.configure("Treeview.Heading", background="lightgrey", foreground="black")
        self.style.map("Treeview", background=[('selected', 'lightyellow')], foreground=[('selected', 'black')])

    def open_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        self.df = pd.read_excel(file_path)
        columns = self.df.columns.tolist()
        self.filter_col_dropdown['values'] = columns
        self.filter_col_var.set(columns[0])

    def preview_results(self):
        if self.df is None:
            messagebox.showerror("Chyba", "Najprv vyber súbor!")
            return

        self.filter_col = self.filter_col_var.get()
        try:
            self.filter_value_num = float(self.filter_val_entry.get())
        except ValueError:
            messagebox.showerror("Chyba", "Zadaj platnú číselnú hodnotu!")
            return

        df_col = pd.to_numeric(
            self.df[self.filter_col].astype(str)
            .str.replace(",", ".")
            .str.replace("kg", "")
            .str.strip(),
            errors='coerce'
        )

        self.filtered_df = self.df[df_col >= self.filter_value_num].copy()
        if self.filtered_df.empty:
            messagebox.showinfo("Výsledok", "Žiadne záznamy nespĺňajú podmienku.")
            return

        filter_index = self.filtered_df.columns.get_loc(self.filter_col)
        self.filtered_df.insert(filter_index + 1, "Správna hmotnosť", [""] * len(self.filtered_df))
        self.row_colors = {}

        if self.tree:
            self.tree.destroy()

        self.tree = ttk.Treeview(self.tree_frame, columns=self.filtered_df.columns.tolist(), show='headings')
        self.tree.pack(fill=tk.BOTH, expand=True)

        for col in self.filtered_df.columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor='center')

        # Tagy
        self.tree.tag_configure('default', background='lightyellow', foreground='black')
        self.tree.tag_configure('editing', background='gold', foreground='black')
        self.tree.tag_configure('greenrow', background='lightgreen', foreground='black')

        # Naplnenie dát
        for i, row in self.filtered_df.iterrows():
            values = list(row)
            item = self.tree.insert("", tk.END, values=values, tags=('default',))
            self.row_colors[item] = 'default'

        self.tree.bind("<Double-1>", self.on_double_click)

    def on_double_click(self, event):
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return

        if self.current_edit_id:
            self.tree.item(self.current_edit_id, tags=(self.row_colors[self.current_edit_id],))
        self.tree.item(item_id, tags=('editing',))
        self.current_edit_id = item_id

        index = int(self.tree.index(item_id))
        row_data = self.filtered_df.iloc[index]

        edit_window = tk.Toplevel(self.master)
        edit_window.title("Úprava riadku")

        entries = {}
        for col in self.filtered_df.columns:
            frame = tk.Frame(edit_window)
            frame.pack(fill=tk.X, padx=5, pady=2)

            tk.Label(frame, text=col, width=20, anchor='w').pack(side=tk.LEFT)
            entry = tk.Entry(frame)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            entry.insert(0, row_data[col])
            if col == "Správna hmotnosť":
                entry.config(bg='lightgreen', fg='black')

            entries[col] = entry

        def save_edit():
            for col, entry in entries.items():
                self.filtered_df.at[self.filtered_df.index[index], col] = entry.get()
            values = list(self.filtered_df.iloc[index])
            self.tree.item(item_id, values=values)

            # Nastavenie farby riadku podľa Správnej hmotnosti
            try:
                # Použijeme .replace(',', '.') pre robustnejšiu konverziu
                weight_str = str(self.filtered_df.at[self.filtered_df.index[index], "Správna hmotnosť"]).replace(',', '.')
                weight = float(weight_str)
            except (ValueError, TypeError):
                weight = None

            # if weight is not None and weight < self.filter_value_num:
            #     self.row_colors[item_id] = 'greenrow'
            # else:
            #     self.row_colors[item_id] = 'default'

            self.tree.item(item_id, tags=(self.row_colors[item_id],))
            self.current_edit_id = None
            edit_window.destroy()

        tk.Button(edit_window, text="OK", command=save_edit).pack(pady=5)

    # --- OPRAVENÁ METÓDA pre Export do Excel s ČÍSELNÝM formátovaním (openpyxl) ---
    def save_results(self):
        if self.filtered_df is None or self.filtered_df.empty:
            messagebox.showwarning("Pozor", "Najprv zobraziť náhľad výsledku!")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered"

        filter_col = self.filter_col

        # Hlavičky
        for j, col in enumerate(self.filtered_df.columns, 1):
            cell = ws.cell(row=1, column=j, value=col)
            if col == filter_col:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # červená
            elif col == "Správna hmotnosť":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # zelená
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Dáta
        for i, row in enumerate(self.filtered_df.itertuples(index=False), 2):
            for j, val in enumerate(row, 1):
                col_name = self.filtered_df.columns[j-1]
                
                # Konverzia na číslo pre správne formátovanie v Exceli
                if col_name == filter_col or col_name == "Správna hmotnosť":
                    try:
                        # Konvertujeme hodnotu na float. Ak je tam "," použijeme ju ako desatinný oddeľovač.
                        numeric_val = float(str(val).replace(',', '.').strip())
                        cell = ws.cell(row=i, column=j, value=numeric_val)
                        cell.number_format = '0.00' # Kľúč k číselnému formátu
                    except (ValueError, TypeError):
                        # Ak to nie je číslo (napr. prázdny reťazec alebo chyba konverzie), zapíšeme pôvodnú hodnotu
                        cell = ws.cell(row=i, column=j, value=val)
                else:
                    # Ostatné stĺpce (text, dátum...)
                    cell = ws.cell(row=i, column=j, value=val)

                # Aplikovanie farieb
                if col_name == filter_col:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif col_name == "Správna hmotnosť":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        output_file = "vysledok.xlsx"
        wb.save(output_file)
        messagebox.showinfo("Uložené", f"Výsledky boli úspešne uložené do súboru: {output_file}")
        os.startfile(output_file)

    # --- Skopírovanie do schránky ako Excel tabuľka (ak si odkomentoval win32com) ---
    # Definovanie metódy copy_to_clipboard_excel pre referenciu
    # def copy_to_clipboard_excel(self):
    #     if self.filtered_df is None or self.filtered_df.empty:
    #         messagebox.showwarning("Pozor", "Najprv zobraziť náhľad výsledku!")
    #         return
    # 
    #     pythoncom.CoInitialize()
    #     excel = win32.Dispatch('Excel.Application')
    #     excel.Visible = False
    #     wb = excel.Workbooks.Add()
    #     ws = wb.Worksheets(1)
    # 
    #     filter_col = self.filter_col
    # 
    #     # Hlavičky
    #     for j, col in enumerate(self.filtered_df.columns, 1):
    #         cell = ws.Cells(1, j)
    #         cell.Value = col
    #         cell.Font.Bold = True
    #         if col == filter_col:
    #             cell.Interior.Color = 0xFF0000
    #         elif col == "Správna hmotnosť":
    #             cell.Interior.Color = 0x00FF00
    # 
    #     # Dáta
    #     for i, row in enumerate(self.filtered_df.values, 2):
    #         for j, val in enumerate(row, 1):
    #             cell = ws.Cells(i, j)
    #             cell.Value = val
    #             # --- Tu by bola tiež potrebná konverzia na float pre správne formátovanie ---
    #             if isinstance(val, (int, float)):
    #                 cell.NumberFormat = '0.00'
    #             col_name = self.filtered_df.columns[j-1]
    #             if col_name == filter_col:
    #                 cell.Interior.Color = 0xFFCCCC
    #             elif col_name == "Správna hmotnosť":
    #                 cell.Interior.Color = 0xCCFFCC
    # 
    #     ws.UsedRange.Copy()
    #     wb.Close(False)
    #     excel.Quit()
    #     messagebox.showinfo("Hotovo", "Tabuľka skopírovaná do schránky ako Excel tabuľka!")

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1200x600")
    app = ExcelFilterApp(root)
    root.mainloop()